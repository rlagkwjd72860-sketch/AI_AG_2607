"""openpyxl을 이용한 Excel 문서 생성 스크립트"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
MAIN_COLOR = "2E74B5"

HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=MAIN_COLOR, end_color=MAIN_COLOR, fill_type="solid")
TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color=MAIN_COLOR)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def create_workbook(output_path: str = "output.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 제목
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "샘플 데이터 시트"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 헤더
    headers = ["이름", "부서", "직급", "입사일"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 데이터
    rows_data = [
        ("홍길동", "개발팀", "사원", "2024-03-02"),
        ("김철수", "디자인팀", "대리", "2022-07-15"),
        ("이영희", "기획팀", "과장", "2019-01-10"),
    ]
    for row_idx, row in enumerate(rows_data, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    # 열 너비 자동 조정
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(output_path)
    print(f"엑셀 파일이 생성되었습니다: {output_path}")


if __name__ == "__main__":
    create_workbook()
